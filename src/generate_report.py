"""Generate stock_report.xlsx from aggregated.json."""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

INPUT = Path(__file__).parent / "outputs" / "aggregated.json"
OUTPUT = Path(__file__).parent / "outputs" / "stock_report.xlsx"

RISK_COLORS = {
    "Conservative": "FF92D050",   # green
    "Moderate":     "FFFFFF00",   # yellow
    "Aggressive":   "FFFF6600",   # orange
    "Speculative":  "FFFF0000",   # red
}

COLUMNS = [
    ("Ticker",       "ticker",       10),
    ("ATH Price",    "ath_price",    12),
    ("ATH Date",     "ath_date",     12),
    ("ATL Price",    "atl_price",    12),
    ("ATL Date",     "atl_date",     12),
    ("Days Between", "days_between", 14),
    ("Speed Label",  "speed_label",  16),
    ("Risk Label",   "risk_label",   14),
    ("Analysis",     "analysis",     60),
]

HEADER_FILL = PatternFill("solid", fgColor="FF1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
ALT_ROW_FILL = PatternFill("solid", fgColor="FFF2F2F2")

thin = Side(style="thin", color="FFD0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def build_report(data: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Report"

    # --- header row ---
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # --- data rows ---
    for row_idx, stock in enumerate(data, start=2):
        alt_fill = ALT_ROW_FILL if row_idx % 2 == 0 else None

        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            value = stock.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="center" if key != "analysis" else "left",
                vertical="top",
                wrap_text=True,
            )
            if alt_fill:
                cell.fill = alt_fill

        # --- risk label color override (column 8) ---
        risk_cell = ws.cell(row=row_idx, column=8)
        risk_value = stock.get("risk_label", "")
        hex_color = RISK_COLORS.get(risk_value)
        if hex_color:
            risk_cell.fill = PatternFill("solid", fgColor=hex_color)
            # Use dark font on yellow/orange for readability
            risk_cell.font = Font(
                bold=True,
                color="FF000000" if risk_value in ("Moderate", "Aggressive") else "FFFFFFFF",
            )

        ws.row_dimensions[row_idx].height = 80

    # --- freeze header ---
    ws.freeze_panes = "A2"

    # --- auto-filter ---
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    wb.save(OUTPUT)
    print(f"Saved: {OUTPUT}")


if __name__ == "__main__":
    with open(INPUT) as f:
        data = json.load(f)
    build_report(data)
